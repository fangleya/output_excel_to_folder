import os
import io
import zipfile
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


# -------------------------- 核心处理模块（增强日志回调） --------------------------
def extract_all_images_from_excel(excel_path, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    image_dict = {}
    try:
        wb = load_workbook(excel_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            image_dict[sheet_name] = []
            images = []
            if hasattr(ws, "_images") and ws._images:  # type: ignore
                images.extend(ws._images)  # type: ignore
            if hasattr(ws, "images") and ws.images:  # type: ignore
                images.extend(ws.images)  # type: ignore
            if hasattr(ws, "_drawing") and ws._drawing:  # type: ignore
                if hasattr(ws._drawing, "charts"):  # type: ignore
                    images.extend([img for img in ws._drawing.charts if isinstance(img, XLImage)])  # type: ignore
                if hasattr(ws._drawing, "_images"):  # type: ignore
                    images.extend(ws._drawing._images)  # type: ignore

            for img in images:
                img_data = None
                try:
                    if hasattr(img, "_data"):
                        img_data = img._data()
                    elif hasattr(img, "ref"):
                        if isinstance(img.ref, io.BytesIO):
                            img_data = img.ref.read()
                            img.ref.seek(0)
                        else:
                            img_data = img.ref
                    elif hasattr(img, "_blob"):
                        img_data = img._blob
                except Exception as e:
                    log(f"读取图片数据失败（工作表{sheet_name}）：{e}")
                    continue
                pos = get_image_cell_position(img)
                if img_data:
                    image_dict[sheet_name].append((img_data, pos))
        wb.close()
    except Exception as e:
        log(f"openpyxl解析失败，启用ZIP兜底方案：{e}")
        image_dict = extract_images_from_zip(excel_path, log_callback)
    return image_dict


def extract_images_from_zip(excel_path, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    image_dict = {}
    temp_images = []
    try:
        with zipfile.ZipFile(excel_path, "r") as zf:
            for file_info in zf.infolist():
                if file_info.filename.startswith("xl/media/") and not file_info.is_dir():
                    file_ext = os.path.splitext(file_info.filename)[1].lower()
                    if file_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp"]:
                        with zf.open(file_info) as f:
                            temp_images.append(f.read())
    except Exception as e:
        log(f"ZIP解压失败：{e}")
        return image_dict

    wb = load_workbook(excel_path, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    for sheet_name in sheet_names:
        image_dict[sheet_name] = []

    if temp_images and sheet_names:
        img_per_sheet = len(temp_images) // len(sheet_names)
        remainder = len(temp_images) % len(sheet_names)
        idx = 0
        for i, sheet_name in enumerate(sheet_names):
            count = img_per_sheet + (1 if i < remainder else 0)
            for _ in range(count):
                if idx < len(temp_images):
                    image_dict[sheet_name].append((temp_images[idx], None))
                    idx += 1
    return image_dict


def get_image_cell_position(img):
    try:
        anchor = getattr(img, "anchor", None)
        if not anchor:
            return None
        cell = getattr(anchor, "_from", None) or getattr(anchor, "from_", None)
        if cell:
            col = cell.col + 1
            row = cell.row + 1
            return (col, row)
        return None
    except Exception:
        return None


def process_excel_to_files(excel_path, output_root, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    if not os.path.exists(excel_path):
        log(f"错误：未找到Excel文件 → {excel_path}")
        return False

    try:
        xl = pd.ExcelFile(excel_path)
        sheet_name = xl.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        if df.empty:
            log(f"错误：Excel文件 [{excel_path}] 的第一个工作表无数据")
            return False

        required_columns = ["初评", "评价图1", "评价图2", "评价图3", "评价图4", "评价图5"]
        if not set(required_columns).issubset(df.columns):
            log(f"错误：Excel标题行不符合要求，需要包含：{required_columns}")
            return False

        image_dict = extract_all_images_from_excel(excel_path, log_callback)
        sheet_images = image_dict.get(sheet_name, [])
        cell_image_map = {}
        for img_data, pos in sheet_images:
            if pos:
                col, row = pos
                cell_image_map[(row, col)] = img_data

        os.makedirs(output_root, exist_ok=True)
        data_sequence = 0
        for row_idx, row_data in df.iterrows():
            excel_row_number = row_idx + 2
            data_sequence += 1
            folder_name = str(data_sequence)
            folder_path = os.path.join(output_root, folder_name)
            has_content = False

            content = str(row_data["初评"]).strip() if pd.notna(row_data["初评"]) else ""
            if content:
                has_content = True
                os.makedirs(folder_path, exist_ok=True)
                txt_path = os.path.join(folder_path, f"{folder_name}.txt")
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write(content)
                log(f"  保存文本: {txt_path}")

            row_image_count = 0
            for img_seq in range(1, 6):
                col_name = f"评价图{img_seq}"
                if col_name not in df.columns:
                    continue
                col_idx = df.columns.get_loc(col_name) + 1
                pos = (excel_row_number, col_idx)
                if pos in cell_image_map:
                    row_image_count += 1
                    has_content = True
                    os.makedirs(folder_path, exist_ok=True)
                    img_data = cell_image_map[pos]
                    try:
                        img_stream = io.BytesIO(img_data)
                        with Image.open(img_stream) as img:
                            img_format = img.format.lower() if img.format else "png"
                            img_name = f"{folder_name}-{row_image_count}.{img_format}"
                            img_path = os.path.join(folder_path, img_name)
                            save_kwargs = {}
                            if img_format == "png":
                                save_kwargs["compress_level"] = 0
                            elif img_format in ["jpg", "jpeg"]:
                                save_kwargs["quality"] = 100
                                save_kwargs["subsampling"] = 0
                            img.save(img_path, **save_kwargs)
                            log(f"  保存图片: {img_path}")
                    except Exception as e:
                        log(f"  保存图片失败（文件夹{folder_name}，图片{row_image_count}）: {e}")
            if has_content:
                log(f"  ✓ 已生成文件夹: {folder_path}")
            else:
                log(f"  ✗ 跳过空白数据行: 文件夹{folder_name}（Excel行{excel_row_number}）")
        log(f"处理完成: {excel_path} -> {output_root}")
        return True
    except Exception as e:
        log(f"处理文件 {excel_path} 时发生错误: {str(e)}\n{traceback.format_exc()}")
        return False
