import sys
import os


# -------------------------- 打包专属路径兼容 --------------------------
# 1. 兼容开发/打包环境的资源根路径
def get_base_path():
    if hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


BASE_PATH = get_base_path()

# 把src目录加入Python路径，解决子包导入报错
sys.path.insert(0, BASE_PATH)

# 2. 强制设置PySide6 Qt插件路径，彻底解决打包后窗口不显示/插件缺失
QT_PLUGIN_PATH = os.path.join(BASE_PATH, "PySide6", "plugins")
if os.path.exists(QT_PLUGIN_PATH):
    os.environ["QT_PLUGIN_PATH"] = QT_PLUGIN_PATH
    os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = os.path.join(QT_PLUGIN_PATH, "platforms")

# 3. 把程序目录加入系统PATH，解决DLL加载失败
os.environ["PATH"] = BASE_PATH + os.pathsep + os.environ.get("PATH", "")

# -------------------------- 以下是import代码 --------------------------


import io
import zipfile
import traceback
from pathlib import Path

from utils.resource_manager import get_icon

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QListWidget,
    QLineEdit,
    QPushButton,
    QTextEdit,
    QFileDialog,
    QMessageBox,
    QProgressBar,
    QLabel,
    QAbstractItemView,
    QListWidgetItem,
)
from PySide6.QtCore import Qt, QThread, Signal, QSettings, QMimeData
from PySide6.QtGui import QDragEnterEvent, QDropEvent, QAction, QFont


# -------------------------- 核心处理模块（增强日志回调） --------------------------
def extract_all_images_from_excel(excel_path, log_callback=None):
    """
    兼容所有openpyxl版本：提取Excel中所有工作表的图片
    返回：字典 {工作表名: [(图片二进制数据, 锚点位置), ...]}
    """

    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    image_dict = {}
    try:
        wb = load_workbook(excel_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            image_dict[sheet_name] = []
            images = []
            if hasattr(ws, "_images") and ws._images:
                images.extend(ws._images)
            if hasattr(ws, "images") and ws.images:
                images.extend(ws.images)
            if hasattr(ws, "_drawing") and ws._drawing:
                if hasattr(ws._drawing, "charts"):
                    images.extend([img for img in ws._drawing.charts if isinstance(img, XLImage)])
                if hasattr(ws._drawing, "_images"):
                    images.extend(ws._drawing._images)

            for img in images:
                img_data = None
                try:
                    if hasattr(img, "_data"):
                        img_data = img._data()
                    elif hasattr(img, "ref"):
                        if isinstance(img.ref, io.BytesIO):
                            img_data = img.ref.read()
                            img.ref.seek(0)
                        else:
                            img_data = img.ref
                    elif hasattr(img, "_blob"):
                        img_data = img._blob
                except Exception as e:
                    log(f"读取图片数据失败（工作表{sheet_name}）：{e}")
                    continue

                pos = get_image_cell_position(img)
                if img_data:
                    image_dict[sheet_name].append((img_data, pos))
        wb.close()
    except Exception as e:
        log(f"openpyxl解析失败，启用ZIP兜底方案：{e}")
        image_dict = extract_images_from_zip(excel_path, log_callback)
    return image_dict


def extract_images_from_zip(excel_path, log_callback=None):
    """兜底方案：直接解压Excel提取所有图片，按工作表分配"""

    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    image_dict = {}
    temp_images = []
    try:
        with zipfile.ZipFile(excel_path, "r") as zf:
            for file_info in zf.infolist():
                if file_info.filename.startswith("xl/media/") and not file_info.is_dir():
                    file_ext = os.path.splitext(file_info.filename)[1].lower()
                    if file_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp"]:
                        with zf.open(file_info) as f:
                            temp_images.append(f.read())
    except Exception as e:
        log(f"ZIP解压失败：{e}")
        return image_dict

    wb = load_workbook(excel_path, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    for sheet_name in sheet_names:
        image_dict[sheet_name] = []

    if temp_images and sheet_names:
        img_per_sheet = len(temp_images) // len(sheet_names)
        remainder = len(temp_images) % len(sheet_names)
        idx = 0
        for i, sheet_name in enumerate(sheet_names):
            count = img_per_sheet + (1 if i < remainder else 0)
            for _ in range(count):
                if idx < len(temp_images):
                    image_dict[sheet_name].append((temp_images[idx], None))
                    idx += 1
    return image_dict


def get_image_cell_position(img):
    """解析图片所在单元格位置（兼容所有openpyxl版本）"""
    try:
        anchor = getattr(img, "anchor", None)
        if not anchor:
            return None
        cell = getattr(anchor, "_from", None) or getattr(anchor, "from_", None)
        if cell:
            col = cell.col + 1
            row = cell.row + 1
            return (col, row)
        return None
    except Exception:
        return None


def process_excel_to_files(excel_path, output_root, log_callback=None):
    """
    处理单个Excel文件，按数据序号生成文件夹及内容，图片命名为"文件夹序号-图片序号"
    支持日志回调
    """

    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    if not os.path.exists(excel_path):
        log(f"错误：未找到Excel文件 → {excel_path}")
        return False

    try:
        # 读取第一个工作表，并获取其真实名称
        xl = pd.ExcelFile(excel_path)
        sheet_name = xl.sheet_names[0]
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        if df.empty:
            log(f"错误：Excel文件 [{excel_path}] 的第一个工作表无数据")
            return False

        # 验证标题行（确保包含初评和图片列）
        required_columns = ["初评", "评价图1", "评价图2", "评价图3", "评价图4", "评价图5"]
        if not set(required_columns).issubset(df.columns):
            log(f"错误：Excel标题行不符合要求，需要包含：{required_columns}")
            return False

        # 提取所有图片
        image_dict = extract_all_images_from_excel(excel_path, log_callback)
        sheet_images = image_dict.get(sheet_name, [])

        # 构建图片位置映射 (行, 列) -> 图片数据（行号为Excel中的实际行号，1起始）
        cell_image_map = {}
        for img_data, pos in sheet_images:
            if pos:  # pos格式为(列, 行)
                col, row = pos
                cell_image_map[(row, col)] = img_data

        # 确保输出根目录存在
        os.makedirs(output_root, exist_ok=True)

        # 处理每行数据（使用数据序号，从1开始）
        data_sequence = 0
        for row_idx, row_data in df.iterrows():
            excel_row_number = row_idx + 2  # 标题行是第1行，数据行从第2行开始
            data_sequence += 1
            folder_name = str(data_sequence)
            folder_path = os.path.join(output_root, folder_name)

            has_content = False

            # 处理初评
            content = str(row_data["初评"]).strip() if pd.notna(row_data["初评"]) else ""
            if content:
                has_content = True
                os.makedirs(folder_path, exist_ok=True)
                txt_path = os.path.join(folder_path, f"{folder_name}.txt")
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write(content)
                log(f"  保存文本: {txt_path}")

            # 处理图片（按行内顺序编号）
            row_image_count = 0
            for img_seq in range(1, 6):
                col_name = f"评价图{img_seq}"
                if col_name not in df.columns:
                    continue

                col_idx = df.columns.get_loc(col_name) + 1  # pandas列索引从0开始，转为Excel列号(1起始)
                pos = (excel_row_number, col_idx)

                if pos in cell_image_map:
                    row_image_count += 1
                    has_content = True
                    os.makedirs(folder_path, exist_ok=True)
                    img_data = cell_image_map[pos]

                    try:
                        img_stream = io.BytesIO(img_data)
                        with Image.open(img_stream) as img:
                            img_format = img.format.lower() if img.format else "png"
                            # 核心命名：文件夹序号-图片序号
                            img_name = f"{folder_name}-{row_image_count}.{img_format}"
                            img_path = os.path.join(folder_path, img_name)

                            save_kwargs = {}
                            if img_format == "png":
                                save_kwargs["compress_level"] = 0
                            elif img_format in ["jpg", "jpeg"]:
                                save_kwargs["quality"] = 100
                                save_kwargs["subsampling"] = 0

                            img.save(img_path, **save_kwargs)
                            log(f"  保存图片: {img_path}")
                    except Exception as e:
                        log(f"  保存图片失败（文件夹{folder_name}，图片{row_image_count}）: {e}")

            if has_content:
                log(f"  ✓ 已生成文件夹: {folder_path}")
            else:
                log(f"  ✗ 跳过空白数据行: 文件夹{folder_name}（Excel行{excel_row_number}）")

        log(f"处理完成: {excel_path} -> {output_root}")
        return True
    except Exception as e:
        log(f"处理文件 {excel_path} 时发生错误: {str(e)}\n{traceback.format_exc()}")
        return False


# -------------------------- 工作线程 --------------------------
class ProcessWorker(QThread):
    log_signal = Signal(str)
    progress_signal = Signal(int, int)  # current, total
    finished_signal = Signal(bool, str)  # success, message

    def __init__(self, excel_files, output_base_dir):
        super().__init__()
        self.excel_files = excel_files
        self.output_base_dir = output_base_dir

    def run(self):
        total = len(self.excel_files)
        if total == 0:
            self.finished_signal.emit(False, "没有待处理的Excel文件")
            return

        success_count = 0
        for idx, excel_path in enumerate(self.excel_files, 1):
            file_name = os.path.basename(excel_path)
            self.log_signal.emit(f"\n>>> 开始处理 [{idx}/{total}]: {file_name}")
            # 构建输出子文件夹：输出根目录/Excel文件名（不含扩展名）
            base_name = os.path.splitext(file_name)[0]
            output_dir = os.path.join(self.output_base_dir, base_name)
            self.log_signal.emit(f"    输出目录: {output_dir}")

            success = process_excel_to_files(excel_path, output_dir, log_callback=self.log_signal.emit)
            if success:
                success_count += 1
                self.log_signal.emit(f"<<< 完成处理: {file_name}\n")
            else:
                self.log_signal.emit(f"<<< 处理失败: {file_name}\n")
            self.progress_signal.emit(idx, total)

        final_msg = f"处理完成！成功 {success_count}/{total} 个文件。"
        self.finished_signal.emit(success_count == total, final_msg)


# -------------------------- 主窗口 --------------------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel数据导出工具")
        self.setMinimumSize(800, 600)
        self.setAcceptDrops(True)  # 允许拖拽

        # 初始化设置（保存最后一次输出路径）
        self.settings = QSettings("MyCompany", "ExcelExportTool")
        self.last_output_dir = self.settings.value("last_output_dir", "")
        if not self.last_output_dir or not os.path.exists(self.last_output_dir):
            self.last_output_dir = "C:/"

        # 创建UI
        self.setup_ui()

        # 初始化线程引用
        self.worker = None

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 1. 输出路径区域
        output_layout = QHBoxLayout()
        output_label = QLabel("输出根目录:")
        self.output_line_edit = QLineEdit()
        self.output_line_edit.setText(self.last_output_dir)
        self.output_line_edit.textChanged.connect(self.save_output_dir)
        browse_btn = QPushButton("浏览...")
        browse_btn.clicked.connect(self.browse_output_dir)
        output_layout.addWidget(output_label)
        output_layout.addWidget(self.output_line_edit, 1)
        output_layout.addWidget(browse_btn)
        main_layout.addLayout(output_layout)

        # 2. Excel文件列表区域
        list_label = QLabel("Excel文件列表（支持拖拽添加，选中后按Delete删除）:")
        main_layout.addWidget(list_label)
        self.file_list_widget = QListWidget()
        self.file_list_widget.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.file_list_widget.setDragDropMode(QAbstractItemView.DragDropMode.DropOnly)
        self.file_list_widget.setAcceptDrops(True)
        self.file_list_widget.installEventFilter(self)  # 为了捕获键盘删除
        main_layout.addWidget(self.file_list_widget)

        # 按钮区域
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始运行")
        self.start_btn.clicked.connect(self.start_processing)
        self.clear_btn = QPushButton("清空列表")
        self.clear_btn.clicked.connect(self.clear_file_list)
        self.remove_btn = QPushButton("移除选中")
        self.remove_btn.clicked.connect(self.remove_selected)
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.clear_btn)
        btn_layout.addWidget(self.remove_btn)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

        # 日志区域
        log_label = QLabel("处理日志:")
        main_layout.addWidget(log_label)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 10))
        main_layout.addWidget(self.log_text)

    def browse_output_dir(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择输出根目录", self.output_line_edit.text())
        if dir_path:
            self.output_line_edit.setText(dir_path)

    def save_output_dir(self):
        current_dir = self.output_line_edit.text()
        if current_dir:
            self.settings.setValue("last_output_dir", current_dir)

    def clear_file_list(self):
        self.file_list_widget.clear()

    def remove_selected(self):
        for item in self.file_list_widget.selectedItems():
            self.file_list_widget.takeItem(self.file_list_widget.row(item))

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        for url in urls:
            file_path = url.toLocalFile()
            if file_path and self.is_excel_file(file_path):
                # 避免重复添加
                if not self.is_file_in_list(file_path):
                    self.file_list_widget.addItem(file_path)
        event.acceptProposedAction()

    def is_excel_file(self, path):
        ext = os.path.splitext(path)[1].lower()
        return ext in [".xlsx", ".xls"]

    def is_file_in_list(self, path):
        for i in range(self.file_list_widget.count()):
            if self.file_list_widget.item(i).text() == path:
                return True
        return False

    def keyPressEvent(self, event):
        # 处理删除键移除选中项
        if event.key() == Qt.Key.Key_Delete:
            self.remove_selected()
        else:
            super().keyPressEvent(event)

    def start_processing(self):
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "提示", "已有处理任务正在运行，请等待完成")
            return

        # 获取输出根目录
        output_base = self.output_line_edit.text().strip()
        if not output_base:
            QMessageBox.warning(self, "提示", "请设置输出根目录")
            return
        # 确保输出根目录存在
        try:
            os.makedirs(output_base, exist_ok=True)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法创建输出目录: {e}")
            return

        # 获取文件列表
        excel_files = []
        for i in range(self.file_list_widget.count()):
            file_path = self.file_list_widget.item(i).text()
            if os.path.exists(file_path):
                excel_files.append(file_path)
            else:
                QMessageBox.warning(self, "文件不存在", f"文件不存在，已忽略: {file_path}")

        if not excel_files:
            QMessageBox.warning(self, "提示", "请先添加有效的Excel文件")
            return

        # 清空日志，准备运行
        self.log_text.clear()
        self.log_text.append("开始批量处理...\n")
        self.start_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, len(excel_files))
        self.progress_bar.setValue(0)

        # 创建工作线程
        self.worker = ProcessWorker(excel_files, output_base)
        self.worker.log_signal.connect(self.append_log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_processing_finished)
        self.worker.start()

    def append_log(self, text):
        self.log_text.append(text)
        # 自动滚动到底部
        cursor = self.log_text.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        self.log_text.setTextCursor(cursor)
        QApplication.processEvents()

    def update_progress(self, current, total):
        self.progress_bar.setValue(current)

    def on_processing_finished(self, success, message):
        self.progress_bar.setVisible(False)
        self.start_btn.setEnabled(True)
        self.append_log(f"\n{message}")
        if success:
            QMessageBox.information(self, "完成", message)
        else:
            QMessageBox.warning(self, "处理完成但有错误", message)


def main():
    app = QApplication(sys.argv)
    # app.setStyle(QStyleFactory.create("windows11"))
    # print(QStyleFactory.keys())
    # print(app.style().name())
    # 全局字体配置：解决PySide6中文乱码问题（Windows推荐雅黑）
    # global_font = QFont("Microsoft YaHei")
    # app.setFont(global_font)
    # 设置应用程序图标
    try:
        app_icon = get_icon("my_app.ico")
        if not app_icon.isNull():
            app.setWindowIcon(app_icon)
    except ImportError:
        pass
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


# -------------------------- 程序入口 --------------------------
if __name__ == "__main__":
    main()
